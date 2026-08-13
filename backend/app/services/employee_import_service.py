"""
Bulk-import employee master data from an uploaded .xlsx workbook (Section
5.1 fields — see the header aliases below). Matches existing employees by
Employee Number or Official Email; creates a new User+Employee pair for
anyone not already in the system.

Column headers are matched case-insensitively and with common wording
variants (including the "Employeement Type" typo seen in real exports),
so a sheet doesn't need to match a schema exactly.
"""
import io
from datetime import date, datetime

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import hash_password
from app.models.employee import Employee
from app.models.enums import EmploymentType, OffboardReason, RoleEnum
from app.models.user import User
from app.repositories.employee_repository import EmployeeRepository
from app.repositories.user_repository import UserRepository
from app.schemas.import_result import EmployeeImportResult, ImportRowError
from app.services.audit_service import AuditService
from app.services.employee_service import FULL_ACCESS_ROLES
from app.utils.password_generator import generate_temp_password

# target field -> acceptable header strings (already lowercased/stripped)
HEADER_ALIASES: dict[str, list[str]] = {
    "employee_number": ["employee number", "employee no", "employee no.", "emp no", "emp id"],
    "full_name": ["employee name", "name", "full name"],
    "department": ["department", "dept"],
    "designation": ["designation", "title"],
    "employment_type": ["employeement type", "employment type", "type"],
    "gender": ["gender"],
    "date_of_birth": ["date of birth", "dob", "birth date"],
    "date_of_joining": ["joined on", "joining date", "date of joining"],
    "offboarded_at": ["leaving date", "relieving date", "date of leaving"],
    "official_email": ["official email id", "official email", "email"],
    "personal_email": ["personal email id", "personal email"],
    "mobile_number": ["mobile number", "mobile no", "mobile no.", "phone", "contact number"],
    "personal_address": ["personal address", "address", "home address", "residential address"],
    "blood_group": ["blood group", "blood type"],
    "emergency_contact": ["emergency contact", "emergency number", "emergency phone"],
    "notice_period_days": ["notice period", "notice period days", "notice days"],
    "bank_account_number": ["bank account number", "bank account details", "account number"],
    "bank_ifsc": ["ifsc", "ifsc code", "bank ifsc"],
    "bank_name": ["bank name", "bank"],
    "pf_number": ["pf number", "pf details", "pf no", "pf no."],
    "status_raw": ["status (employee/relieved)", "status", "status(employee/ relieved)"],
    "reporting_manager_name": ["reporting manager", "manager"],
}

EMPLOYMENT_TYPE_MAP = {
    "full time": EmploymentType.FULL_TIME,
    "fulltime": EmploymentType.FULL_TIME,
    "full-time": EmploymentType.FULL_TIME,
    "intern": EmploymentType.INTERN,
    "internship": EmploymentType.INTERN,
    "contract": EmploymentType.CONTRACT,
    "contractor": EmploymentType.CONTRACT,
}


def _normalize_header(raw: str) -> str:
    return " ".join(str(raw).strip().lower().split())


def _cell_to_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%b %d, %Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _cell_to_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


class EmployeeImportService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.employees = EmployeeRepository(db)
        self.users = UserRepository(db)
        self.audit = AuditService(db)

    async def import_from_excel(self, file: UploadFile, requester: User) -> EmployeeImportResult:
        if requester.role not in FULL_ACCESS_ROLES:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only HR Admin, HR Executive, or System Admin can import employees.",
            )

        try:
            raw = await file.read()
            workbook = load_workbook(filename=io.BytesIO(raw), data_only=True)
            sheet = workbook.active
        except Exception as exc:  # noqa: BLE001 — surfaced to the caller either way
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Could not read this file as an Excel workbook: {exc}",
            )

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="The sheet is empty.")

        header_row = rows[0]
        column_map: dict[str, int] = {}
        for idx, header in enumerate(header_row):
            if header is None:
                continue
            normalized = _normalize_header(header)
            for field, aliases in HEADER_ALIASES.items():
                if normalized in aliases and field not in column_map:
                    column_map[field] = idx

        if "official_email" not in column_map and "employee_number" not in column_map:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "Couldn't find an 'Official Email ID' or 'Employee Number' column — "
                    "at least one is needed to match rows to employees."
                ),
            )

        def get(row: tuple, field: str):
            idx = column_map.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        created = updated = skipped = 0
        errors: list[ImportRowError] = []

        for row_num, row in enumerate(rows[1:], start=2):
            if row is None or all(v is None for v in row):
                continue

            identifier = _cell_to_str(get(row, "full_name")) or _cell_to_str(get(row, "official_email")) or f"row {row_num}"
            try:
                employee_number = _cell_to_str(get(row, "employee_number"))
                official_email = _cell_to_str(get(row, "official_email"))

                existing_user = None
                if official_email:
                    existing_user = await self.users.get_by_email(official_email)
                if existing_user is None and employee_number:
                    existing_user = await self.users.get_by_employee_id(employee_number)

                employee = (
                    await self.employees.get_by_user_id(existing_user.id) if existing_user else None
                )

                fields: dict = {}
                if (v := _cell_to_str(get(row, "full_name"))) is not None:
                    fields["full_name"] = v
                if (v := _cell_to_str(get(row, "department"))) is not None:
                    fields["department"] = v
                if (v := _cell_to_str(get(row, "designation"))) is not None:
                    fields["designation"] = v
                if (v := _cell_to_str(get(row, "gender"))) is not None:
                    fields["gender"] = v

                if (v := _cell_to_date(get(row, "date_of_birth"))) is not None:
                    fields["date_of_birth"] = v

                if (v := _cell_to_str(get(row, "personal_address"))) is not None:
                    fields["personal_address"] = v

                if (v := _cell_to_str(get(row, "blood_group"))) is not None:
                    fields["blood_group"] = v

                if (v := _cell_to_str(get(row, "emergency_contact"))) is not None:
                    fields["emergency_contact"] = v

                notice_raw = get(row, "notice_period_days")
                if notice_raw is not None and str(notice_raw).strip() != "":
                    try:
                        fields["notice_period_days"] = int(float(str(notice_raw).strip()))
                    except (ValueError, TypeError):
                        pass
                if (v := _cell_to_str(get(row, "personal_email"))) is not None:
                    fields["personal_email"] = v
                if (v := _cell_to_str(get(row, "mobile_number"))) is not None:
                    fields["mobile_number"] = v
                if (v := _cell_to_str(get(row, "bank_account_number"))) is not None:
                    fields["bank_account_number"] = v
                if (v := _cell_to_str(get(row, "bank_ifsc"))) is not None:
                    fields["bank_ifsc"] = v
                if (v := _cell_to_str(get(row, "bank_name"))) is not None:
                    fields["bank_name"] = v
                if (v := _cell_to_str(get(row, "pf_number"))) is not None:
                    fields["pf_number"] = v
                if (v := _cell_to_date(get(row, "date_of_joining"))) is not None:
                    fields["date_of_joining"] = v

                emp_type_raw = _cell_to_str(get(row, "employment_type"))
                if emp_type_raw and emp_type_raw.lower() in EMPLOYMENT_TYPE_MAP:
                    fields["employment_type"] = EMPLOYMENT_TYPE_MAP[emp_type_raw.lower()]

                # Status / offboarding — only acted on if the sheet clearly
                # says someone has left; otherwise leave status untouched.
                status_raw = (_cell_to_str(get(row, "status_raw")) or "").lower()
                leaving_date = _cell_to_date(get(row, "offboarded_at"))
                if "relieved" in status_raw or "resign" in status_raw or "terminat" in status_raw or leaving_date:
                    fields["status"] = "offboarded"
                    fields["offboard_reason"] = (
                        OffboardReason.TERMINATED if "terminat" in status_raw else OffboardReason.RESIGNED
                    )
                    fields["offboarded_at"] = leaving_date or date.today()

                manager_name = _cell_to_str(get(row, "reporting_manager_name"))
                manager_id = None
                if manager_name:
                    all_employees = await self.employees.list_all_for_aggregation()
                    match = next(
                        (e for e in all_employees if e.full_name.strip().lower() == manager_name.strip().lower()),
                        None,
                    )
                    if match:
                        manager_id = match.id
                    # Silently skip if no match — best-effort only, not an error.

                if employee is not None:
                    # ---- Update existing employee ----
                    for key, value in fields.items():
                        setattr(employee, key, value)
                    if manager_id is not None:
                        employee.reporting_manager_id = manager_id
                    await self.employees.save(employee)
                    updated += 1
                else:
                    # ---- Create new employee ----
                    if not official_email:
                        errors.append(
                            ImportRowError(
                                row=row_num,
                                identifier=identifier,
                                error="No existing employee matched, and no Official Email ID was given to create a new one.",
                            )
                        )
                        skipped += 1
                        continue
                    if existing_user is not None:
                        errors.append(
                            ImportRowError(
                                row=row_num,
                                identifier=identifier,
                                error="A user with this email/employee number already exists but has no employee profile — skipped rather than guessing.",
                            )
                        )
                        skipped += 1
                        continue

                    temp_password = generate_temp_password()
                    user = User(
                        official_email=official_email,
                        employee_id=employee_number,
                        hashed_password=hash_password(temp_password),
                        role=RoleEnum.EMPLOYEE,
                        is_active=True,
                    )
                    await self.users.create(user)

                    new_employee = Employee(
                        user_id=user.id,
                        full_name=fields.get("full_name", identifier),
                        reporting_manager_id=manager_id,
                        **{k: v for k, v in fields.items() if k != "full_name"},
                    )
                    await self.employees.create(new_employee)
                    created += 1

            except Exception as exc:  # noqa: BLE001 — one bad row shouldn't kill the whole import
                errors.append(ImportRowError(row=row_num, identifier=identifier, error=str(exc)))
                skipped += 1

        await self.audit.log(
            requester.id,
            "employee_bulk_import",
            "employee",
            "bulk",
            detail=f"created={created} updated={updated} skipped={skipped}",
        )

        return EmployeeImportResult(
            total_rows=len(rows) - 1,
            created=created,
            updated=updated,
            skipped=skipped,
            errors=errors,
        )
